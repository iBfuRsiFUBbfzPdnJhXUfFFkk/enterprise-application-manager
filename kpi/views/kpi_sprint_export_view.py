from io import BytesIO

from django.http import HttpRequest, HttpResponse
from openpyxl.styles import Font, Side, PatternFill, Border
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook

from core.models.sprint import Sprint
from kpi.models.key_performance_indicator_sprint import KeyPerformanceIndicatorSprint


def kpi_sprint_export_view(request: HttpRequest, uuid: str) -> HttpResponse:
    sprint = Sprint.from_uuid(uuid=uuid)
    sprint_kpis = KeyPerformanceIndicatorSprint.objects.filter(sprint=sprint)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = f"Sprint {sprint.name} KPIs"

    # Define styling
    header_font = Font(bold=True, color="FFFFFF")
    group_header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    subheader_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"),
                         bottom=Side(style="thin"))
    thick_border = Border(left=Side(style="thick"), right=Side(style="thick"), top=Side(style="thick"),
                          bottom=Side(style="thick"))

    # Updated main group headers with column ranges (reflecting the new context switching column)
    group_headers = [
        ("Developer", 1, 1),
        ("Availability", 2, 5),
        ("Commitments", 6, 7),
        ("Code Review Activity", 8, 10),
        ("Development Output", 11, 13),  # Increased to include Context Switching
        ("Performance Metrics", 14, 15),
    ]

    # Create header rows
    sheet.append([""] * 15)  # Main group headers row
    sheet.append(
        [
            "Developer",
            "Base Capacity",
            "Holidays",
            "PTO Days",
            "Adjusted Capacity",
            "Committed",
            "Delivered",
            "Reviews",
            "Comments",
            "Threads",
            "Issues",
            "Code Changes",
            "Context Switching",  # Added new column
            "Velocity",
            "Accuracy",
        ]
    )

    # Add main group headers with borders
    for title, start_col, end_col in group_headers:
        cell = sheet.cell(row=1, column=start_col, value=title)
        cell.font = header_font
        cell.fill = group_header_fill
        cell.border = thick_border  # Add bold outer border

        if start_col != end_col:
            sheet.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
            # Apply border to merged cells
            for col in range(start_col, end_col + 1):
                sheet.cell(row=1, column=col).border = thick_border

    # Style subheaders
    for col in range(1, 16):  # Adjusted to 16 to account for the new column
        cell = sheet.cell(row=2, column=col)
        cell.font = Font(bold=True)
        cell.fill = subheader_fill
        cell.border = thin_border
        sheet.column_dimensions[get_column_letter(col)].width = 15

    for kpi in sprint_kpis:
        code_changes = "---" if kpi.coerced_number_of_code_lines_added == 0 and kpi.coerced_number_of_code_lines_removed == 0 else f"+{kpi.coerced_number_of_code_lines_added}/-{kpi.coerced_number_of_code_lines_removed} lines"
        person = kpi.person_developer
        sprint = kpi.sprint

        row = [
            person.gitlab_sync_username if person else "N/A",
            kpi.coerced_scrum_capacity_base,
            sprint.number_of_holidays_during_sprint if sprint else 0,
            kpi.coerced_number_of_paid_time_off_days,
            kpi.adjusted_capacity,
            kpi.coerced_number_of_story_points_committed_to,
            kpi.coerced_number_of_story_points_delivered,
            kpi.coerced_number_of_merge_requests_approved,
            kpi.coerced_number_of_comments_made,
            kpi.coerced_number_of_threads_made,
            kpi.coerced_number_of_issues_written,
            code_changes,
            kpi.coerced_number_of_context_switches,  # 🆕 New Context Switching column
            kpi.capacity_based_velocity,
            kpi.commitment_accuracy,
        ]

        sheet.append(row)

    # Apply data formatting
    for row in sheet.iter_rows(min_row=3):
        for cell in row:
            cell.border = thin_border
            if cell.column_letter in ["N", "O"]:  # Performance metrics columns (Velocity & Accuracy)
                cell.number_format = "0.00"

    # Save to response
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    response = HttpResponse(output.getvalue(),request=request,
                            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="sprint_{sprint.name}_kpis.xlsx"'
    return response
